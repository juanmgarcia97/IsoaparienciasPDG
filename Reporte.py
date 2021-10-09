#Al tratarse de una librería externa, lo primero que debes hacer para usar openpyxl es instalarla. 
# Para ello es necesario usar el siguiente comando: pip install openpyxl

#pip es un paquete como cualquier otro, la forma más segura y legible de usarlo y de asegurarte 
# de instalar en el intérprete que quieres instalar es ejecutarlo como módulo de la forma: py -m pip install openpyxl


#Debemo improtar openpyxl
import openpyxl
#Para crear un nuevl libro usamos el comando
wb = openpyxl.Workbook()
#para acceder a la hoja activa usamos el atributo "active"
hoja = wb.active
hoja.title="PRUEBA"
print(f'Hoja activa: {wb.active.title}')


# Añade la hoja 'Hoja' al final (por defecto)
#>>> hoja1 = wb.create_sheet("Hoja")

# Añade la hoja 'Hoja' en la primera posición. Como el nombre
# 'Hoja' ya existe, le añade el número 1 al final del nombre
#>>> hoja2 = wb.create_sheet("Hoja", 0)

# Añade la hoja 'Otra hoja' en la posición 1
#>>> wb.create_sheet(index=1, title="Otra hoja")
wb.create_sheet(index=1, title="HojaPrueba")
wb.create_sheet(index=2, title="HojaPrueba2")
# Muestra los nombres de las hojas
#>>> print(wb.sheetnames)
#['Hoja1', 'Otra hoja', 'Valores', 'Hoja']

#Para copiar una hoja podemos usar el siguiente codigo
#>>> origen = wb.active
#>>> nueva = wb.copy_worksheet(origen)

#Para acceder a otra hoja podemos usar el nombre y tambien podemos volver hoja activa
#hoja = wb['Otra hoja']
#>>> wb.active = hoja


#podemos acceder a todas las hojas e iterar sobre ellas con el sigueinte ejemplo:
#>>> for hoja in wb:
# ...     print(hoja.title)
#Hoja1
#Otra hoja
#Valores
#Hoja


#ejemplo de como acceder a una celda
#>>> wb = openpyxl.Workbook()
#>>> hoja = wb.active
#>>> a1 = hoja["A1"]
#>>> print(a1.value)
#None

#otra forma de accerder a celdas:
#b2 = hoja.cell(row=2, column=2)
#b1 = hoja.cell(row=1, column=2, value=20)
#c1 = hoja.cell(row=1, column=3)
#c1.value = 30


#Para asignar valores a celdas podemos usar lo siguiente:
#hoja["A1"] = 10

#Para guardar un archivo excel usamos el metodo save
wb.save('reportePRuebas.xlsx')













